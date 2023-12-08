#!/usr/bin/env python3
# -*- coding: utf-8 -*-

from sys import argv
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font

def translate(*sources):
    import json
    from volcengine.ApiInfo import ApiInfo
    from volcengine.Credentials import Credentials
    from volcengine.ServiceInfo import ServiceInfo
    from volcengine.base.Service import Service

    k_access_key = 'AKLTZTU1MGIzYzVkNWU4NGExNWExNzJlZTc5ZGE2ZTdjODA' # https://console.volcengine.com/iam/keymanage/
    k_secret_key = 'TUdGbU4yUTRNV0ZrTkRNNE5ETmhORGcwWlRobU1qa3dZMll5WXpGbFltWQ=='
    k_service_info = \
        ServiceInfo('open.volcengineapi.com',
                    {'Content-Type': 'application/json'},
                    Credentials(k_access_key, k_secret_key, 'translate', 'cn-north-1'),
                    5,
                    5)
    k_query = {
        'Action': 'TranslateText',
        'Version': '2020-06-01'
    }
    k_api_info = {
        'translate': ApiInfo('POST', '/', k_query, {}, {})
    }
    service = Service(k_service_info, k_api_info)
    body = {
        'TargetLanguage': 'zh',
        'TextList': sources,
    }
    res = service.json('translate', {}, json.dumps(body))
    r = json.loads(res)
    # print(r)
    re = []
    for t in r['TranslationList']:
        re.append(t['Translation'])
    return tuple(re)

def process_file(file_name, sheet_name):
    wb1 = load_workbook(file_name)
    ws1 = wb1[sheet_name]
    wb2 = Workbook()
    wb2.remove(wb2.active)
    i = 0
    last_unit = 0
    words = {}
    for row in ws1.iter_rows(min_row=2, min_col=1, max_col=6):
        word = row[0].value
        unit = row[1].value
        if unit != last_unit:
            words[unit] = []
            last_unit = unit
        words[unit].append(word)
    # print(words)
    i = 0
    last_unit = 0
    translations = []
    for row in ws1.iter_rows(min_row=2, min_col=1, max_col=6):
        word = row[0].value
        unit = row[1].value
        defination = row[3].value
        pos = row[4].value
        example = row[5].value
        if unit != last_unit:
            r = ['NO', "WORD", "PoS", "DEFINITION & EXAMPLE", "UNIT", "TRANSLATION"]
            ws2 = wb2.create_sheet("UNIT %s" % unit)
            ws2.append(r)
            # change style
            col = ws2.column_dimensions['D']
            col.alignment = Alignment(vertical="top", wrapText=True)
            col.width = 80
            col = ws2.column_dimensions['B']
            col.alignment = Alignment(horizontal="center", vertical="top", wrapText=True)
            col.width = 20
            col = ws2.column_dimensions['A']
            col.alignment = Alignment(vertical="top", wrapText=True)
            col.width = 5
            col = ws2.column_dimensions['E']
            col.alignment = Alignment(vertical="top", wrapText=True)
            col.width = 5
            col = ws2.column_dimensions['C']
            col.alignment = Alignment(vertical="top", wrapText=True)
            col.width = 10
            col = ws2.column_dimensions['F']
            col.alignment = Alignment(vertical="top", wrapText=True)
            col.width = 10
            last_unit = unit
            i = 1
            translations = translate(*words[unit])
        # print(translations)
        r = [i, word, pos, "%s\r%s" % (defination, example), unit, translations[i-1]]
        ws2.append(r)
        i = i + 1
    wb2.save("think 1.xlsx")

if __name__ == "__main__":
    import sys
    if len(sys.argv) < 2:
        print("Need file name", file=sys.stderr)
        sys.exit(1)
    sheet_name = 'THINK L1 FRENCH' if len(sys.argv) <= 2 else sys.argv[2]
    process_file(sys.argv[1], sheet_name)
